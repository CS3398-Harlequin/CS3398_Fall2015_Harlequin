import openpyxl

#wb = openpyxl.load_workbook(filename='austinTXtwitterdata.xlsx', read_only=True)
#ws = wb['results']

def getSheetValue(workbookname, worksheetname, column, row):
    wb = openpyxl.load_workbook(filename=workbookname, read_only=True)
    ws = wb[worksheetname]
    return ws.cell(column=column, row=row).value


#should format that output better
print((getSheetValue('austinTXtwitterdata.xlsx', 'results', 4, 3)).encode('utf-8', 'ignore'))

